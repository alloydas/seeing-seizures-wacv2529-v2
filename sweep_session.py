"""
Full-session seizure localization sweep (video).

Slides the windowed R(2+1)D-18 classifier (train_video_windows.py) across an
UNCUT session recording and emits a P(seizure) trace over wall-clock time, then
thresholds it into detected intervals and scores them against the .xlsx
annotations.

This is deliberately NOT run on the cut clips: every seizure clip in Data_* was
cut with Pre-buffer = 10.0 s exactly, so onset always sits at t=10 s and any
within-clip onset metric measures the cutter, not the model.

The sweep decodes each video ONCE, sequentially. Frames the windows do not need
are skipped with grab() (parse, no decode); only the ~16-per-window sampled
frames are retrieve()'d. Seeking per window would be ruinous on copy-mode mp4s.

Annotation/timestamp helpers are copied from cut_seizure_clips.py rather than
imported -- that module calls parse_args() at import time. Keep in sync.

Usage:
    python sweep_session.py --session "/media/.../RN197/11-25-2023" \
        --ckpt video_win_out_W4/last.pt --subject RN197 --out sweep_out/RN197_11-25
"""
import argparse, glob, json, os, re, sys
from datetime import datetime, timedelta
from pathlib import Path

import cv2, numpy as np, torch
from train_classifier import build_model, KINETICS_MEAN, KINETICS_STD

# Per-camera cage geometry, mirroring sh/_crop_all.sh. Getting the side wrong
# silently poisons the sweep, so this table is explicit rather than inferred.
#   subject: (left_frac, x_frac, top_frac, y_frac)
CROPS = {
    "RN197": (0.60, 0.00, 0.80, 0.00),   # side-by-side left  (pairs with RN213)
    "RN213": (0.40, 0.60, 0.80, 0.00),   # side-by-side right
    "RN222": (0.60, 0.00, 0.80, 0.00),
    "RN237": (0.60, 0.00, 0.80, 0.00),
    "RN204": (0.40, 0.60, 0.80, 0.00),
    "RN235": (0.40, 0.60, 0.80, 0.00),
    "RN238": (0.40, 0.60, 0.80, 0.00),
    "RN199": (0.50, 0.00, 0.60, 0.30),   # 2x2 lower-left
    "RN245": (0.50, 0.00, 0.60, 0.30),
    "RN216": (0.48, 0.52, 0.60, 0.30),   # 2x2 lower-right
    # RoomD camera (RN242-RN243), 2x2 grid, RN243 = lower-right. NOTE the source
    # table sh/_crop_roomd.sh stores "RN243 0.50 0.40 0.45 0.37" in a DIFFERENT
    # field order than sh/_crop_all.sh: RoomD parses `s xf wf yf hf` while RoomC
    # parses `s wf xf hf yf`. Transposed to this dict's (left, x, top, y) order
    # and verified visually against a decoded frame on 2026-08-11.
    "RN243": (0.40, 0.50, 0.37, 0.45),
}


def parse_args():
    p = argparse.ArgumentParser(description="Sweep a full session for seizure intervals.")
    p.add_argument("--session", required=True, help="Raw session folder (.mp4 + .XML + .xlsx)")
    p.add_argument("--ckpt", required=True, help="Windowed model checkpoint (train_video_windows.py)")
    p.add_argument("--subject", required=True, choices=sorted(CROPS), help="Selects cage crop")
    p.add_argument("--arch", default="r2plus1d_win",
                   help="r2plus1d_win (train_video_windows.py ckpt) or a train_pooled.py "
                        "backbone name: r2plus1d|mvit|mvit_v1|swin|swin_s|s3d|x3d|slowfast")
    p.add_argument("--out", required=True, help="Output folder for trace + results")
    p.add_argument("--win", type=float, default=4.0)
    p.add_argument("--stride", type=float, default=2.0)
    p.add_argument("--frames", type=int, default=16)
    p.add_argument("--size", type=int, default=112)
    p.add_argument("--batch_size", type=int, default=64)
    p.add_argument("--device", default="cuda:0")
    p.add_argument("--max_videos", type=int, default=None, help="Cap videos swept (smoke-test)")
    p.add_argument("--max_seconds", type=float, default=None, help="Cap seconds per video (smoke-test)")
    return p.parse_args()


# ── timestamp + annotation parsing (copied from cut_seizure_clips.py) ─────────

def _xml_start_time(fpath):
    try:
        from xml.etree import ElementTree as ET
        base = os.path.splitext(fpath)[0]
        for ext in (".XML", ".xml"):
            if os.path.exists(base + ext):
                root = ET.parse(base + ext).getroot()
                node = root.find("DSI_utc_start_time")
                if node is not None and node.text:
                    return datetime.fromtimestamp(int(node.text.strip()))
                return None
    except Exception:
        pass
    return None


def parse_video_files(folder):
    p14, p12 = re.compile(r"(\d{14})"), re.compile(r"(\d{12})")
    videos = []
    for fpath in sorted(glob.glob(os.path.join(folder, "*.mp4"))):
        start_dt = _xml_start_time(fpath)
        if start_dt is None:
            for pat, fmt in [(p14, "%Y%m%d%H%M%S"), (p12, "%Y%m%d%H%M")]:
                m = pat.search(os.path.basename(fpath))
                if m:
                    start_dt = datetime.strptime(m.group(1), fmt)
                    break
        if start_dt is None:
            print(f"  [warn] no timestamp for {os.path.basename(fpath)} -- skipped")
            continue
        cap = cv2.VideoCapture(fpath)
        fps = cap.get(cv2.CAP_PROP_FPS) or 15.0
        n = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        cap.release()
        if n <= 0:
            print(f"  [warn] unreadable {os.path.basename(fpath)} -- skipped")
            continue
        videos.append(dict(path=fpath, start=start_dt, fps=fps, n=n,
                           end=start_dt + timedelta(seconds=n / fps)))
    return videos


def parse_annotations(excel_path):
    import openpyxl
    ws = openpyxl.load_workbook(excel_path).active
    headers = [str(c.value).strip().lower() if c.value else ""
               for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(*names):
        for n in names:
            for i, h in enumerate(headers):
                if n in h:
                    return i
        return None

    i_label, i_dur = col("label", "stage", "type"), col("dur", "length")
    i_start, i_end = col("start"), col("end")
    if i_start is None:
        sys.exit(f"[ERROR] no 'Start' column. Headers: {headers}")

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        start = row[i_start]
        if start is None or not hasattr(start, "hour"):
            continue                      # skips the '[]' units row
        try:
            dur = float(row[i_dur]) if i_dur is not None and row[i_dur] is not None else 30.0
        except (ValueError, TypeError):
            dur = 30.0
        end = row[i_end] if i_end is not None else None
        if not hasattr(end, "hour"):
            end = start + timedelta(seconds=dur)
        label = str(row[i_label]).strip() if i_label is not None and row[i_label] else "Unknown"
        out.append(dict(label=label, start=start, end=end, duration=dur))
    return out


# ── sweep ────────────────────────────────────────────────────────────────────

def crop_box(w, h, fr):
    """Reproduce crop_clips.py's ffmpeg filter: trunc(dim*frac/2)*2."""
    left, xf, top, yf = fr
    cw = int(w * left / 2) * 2
    ch = int(h * top / 2) * 2
    cx = int(w * xf / 2) * 2
    cy = int(h * yf / 2) * 2
    return cx, cy, cw, ch


def sweep_video(vid, model, dev, fr, args):
    """Yield (t0_rel_seconds, prob) for every window in one video."""
    cap = cv2.VideoCapture(vid["path"])
    fps, nF = vid["fps"], vid["n"]
    if args.max_seconds:
        nF = min(nF, int(args.max_seconds * fps))
    win_f = int(round(args.win * fps))
    stride_f = int(round(args.stride * fps))
    box = None

    active, pend_t, pend_x, out = [], [], [], []
    next_w0 = 0

    def flush():
        if not pend_x:
            return
        x = torch.from_numpy(np.stack(pend_x)).to(dev, non_blocking=True)
        with torch.no_grad(), torch.autocast("cuda", dtype=torch.float16):
            p = torch.softmax(model(x).float(), 1)[:, 1]
        out.extend(zip(pend_t, p.cpu().numpy().tolist()))
        pend_t.clear(); pend_x.clear()

    for i in range(nF):
        if not cap.grab():
            break
        if i == next_w0 and i + win_f - 1 < nF:
            idx = np.linspace(i, i + win_f - 1, args.frames).round().astype(int)
            m = {}
            for slot, f in enumerate(idx):
                m.setdefault(int(f), []).append(slot)
            active.append(dict(f0=i, f1=i + win_f - 1, m=m,
                               buf=[None] * args.frames))
            next_w0 += stride_f
        if any(i in w["m"] for w in active):
            ok, frame = cap.retrieve()
            if ok:
                if box is None:
                    box = crop_box(frame.shape[1], frame.shape[0], fr)
                cx, cy, cw, ch = box
                f = frame[cy:cy + ch, cx:cx + cw]
                f = cv2.resize(cv2.cvtColor(f, cv2.COLOR_BGR2RGB),
                               (args.size, args.size), interpolation=cv2.INTER_AREA)
                for w in active:
                    for slot in w["m"].get(i, []):
                        w["buf"][slot] = f
        for w in [w for w in active if w["f1"] == i]:
            if any(b is None for b in w["buf"]):
                continue                   # decode gap -- drop rather than zero-fill
            a = np.stack(w["buf"]).astype(np.float32) / 255.0
            a = ((a - KINETICS_MEAN) / KINETICS_STD).transpose(3, 0, 1, 2)
            pend_t.append(w["f0"] / fps); pend_x.append(a)
            if len(pend_x) >= args.batch_size:
                flush()
        active = [w for w in active if w["f1"] > i]
    flush()
    cap.release()
    return out


# ── detection + scoring ──────────────────────────────────────────────────────

def extract_intervals(times, probs, thr, win, min_dur=4.0, merge_gap=4.0):
    """Threshold the trace into contiguous detected intervals (seconds)."""
    hits = [(t, t + win) for t, p in zip(times, probs) if p >= thr]
    if not hits:
        return []
    merged = [list(hits[0])]
    for s, e in hits[1:]:
        if s - merged[-1][1] <= merge_gap:
            merged[-1][1] = max(merged[-1][1], e)
        else:
            merged.append([s, e])
    return [tuple(m) for m in merged if m[1] - m[0] >= min_dur]


def match(det, gt, min_iou=0.3):
    """Greedy one-to-one match of detections to ground truth by temporal IoU."""
    pairs, used = [], set()
    for gi, (gs, ge) in enumerate(gt):
        best, bi = 0.0, None
        for di, (ds, de) in enumerate(det):
            if di in used:
                continue
            inter = max(0.0, min(ge, de) - max(gs, ds))
            union = max(ge, de) - min(gs, ds)
            iou = inter / union if union > 0 else 0.0
            if iou > best:
                best, bi = iou, di
        if bi is not None and best >= min_iou:
            used.add(bi); pairs.append((gi, bi, best))
    return pairs, used


def main():
    args = parse_args()
    out = Path(args.out); out.mkdir(parents=True, exist_ok=True)
    dev = torch.device(args.device if torch.cuda.is_available() else "cpu")
    fr = CROPS[args.subject]

    vids = parse_video_files(args.session)
    if args.max_videos:
        vids = vids[:args.max_videos]
    xl = glob.glob(os.path.join(args.session, "*.xlsx"))
    if not xl:
        sys.exit(f"[ERROR] no .xlsx in {args.session}")
    seiz = parse_annotations(xl[0])

    print(f"session   : {args.session}")
    print(f"subject   : {args.subject}  crop={fr}")
    print(f"videos    : {len(vids)}  ({sum(v['n']/v['fps'] for v in vids)/3600:.1f} h)")
    print(f"annotated : {len(seiz)} seizures")

    ck = torch.load(args.ckpt, map_location="cpu")
    if args.arch == "r2plus1d_win":
        # the windowed model from train_video_windows.py, built by train_classifier
        model = build_model(2)
    else:
        # whole-clip detection backbones from train_pooled.py. NOTE these were trained
        # on FULL clips (16 frames spread over the whole ~40-80 s clip), so applying
        # them to a 4 s window is a temporal-extent mismatch: they see the same 16
        # frames over a 10-20x shorter span. Only r2plus1d_win was trained at this
        # window, so cross-backbone differences below partly reflect that, not backbone
        # quality. Kept identical across backbones so at least they are comparable to
        # each other.
        from train_pooled import build_video_model
        built = build_video_model(args.arch, 2, pretrained=False)   # NOT `out`:
        # `out` is already the output-directory Path in this scope, and shadowing it
        # made np.savez(out / "trace.npz") fail with "unsupported operand type(s)
        # for /: 'tuple' and 'str'" after a full sweep had already been computed.
        model, df, ds = built if isinstance(built, tuple) else (built, None, None)
        # Each backbone has its own expected input geometry and its checkpoint was
        # trained at it: mvit/mvit_v1/x3d/s3d 16x224, slowfast/swin/swin_s 32x224,
        # r2plus1d 16x112. Feeding the r2plus1d default (16x112) to the others failed
        # deep inside torchvision (avg_pool3d kernel larger than input, pos-embedding
        # size mismatch), so honour the per-arch defaults.
        if df:
            args.frames = df
        if ds:
            args.size = ds
        print(f"arch      : {args.arch}  frames={args.frames} size={args.size}")
    model.load_state_dict(ck["model"])
    model.to(dev).eval()
    print(f"ckpt      : {args.ckpt} (epoch {ck.get('epoch')}, win {ck.get('win')})\n", flush=True)

    T, P, V = [], [], []
    for k, v in enumerate(vids):
        r = sweep_video(v, model, dev, fr, args)
        for t, p in r:
            T.append(v["start"].timestamp() + t); P.append(p); V.append(k)
        print(f"  [{k+1}/{len(vids)}] {os.path.basename(v['path'])}: "
              f"{len(r)} windows", flush=True)

    T, P = np.array(T), np.array(P, np.float32)
    np.savez(out / "trace.npz", t_abs=T, prob=P, vid=np.array(V),
             win=args.win, stride=args.stride)

    gt = [(s["start"].timestamp(), s["end"].timestamp()) for s in seiz]
    hours = sum(v["n"] / v["fps"] for v in vids) / 3600.0
    rows = []
    for thr in (0.5, 0.7, 0.9, 0.95, 0.99):
        det = extract_intervals(T, P, thr, args.win)
        pairs, used = match(det, gt)
        tp, fp, fn = len(pairs), len(det) - len(used), len(gt) - len(pairs)
        on = [abs(det[di][0] - gt[gi][0]) for gi, di, _ in pairs]
        off = [abs(det[di][1] - gt[gi][1]) for gi, di, _ in pairs]
        rows.append(dict(thr=thr, n_det=len(det), tp=tp, fp=fp, fn=fn,
                         recall=tp / len(gt) if gt else 0.0,
                         precision=tp / len(det) if det else 0.0,
                         fp_per_hour=fp / hours if hours else 0.0,
                         onset_mae=float(np.mean(on)) if on else None,
                         offset_mae=float(np.mean(off)) if off else None,
                         mean_iou=float(np.mean([p[2] for p in pairs])) if pairs else None))

    print(f"\n=== localization vs {len(gt)} annotated seizures over {hours:.1f} h ===")
    print(f"  {'thr':>5} {'det':>6} {'TP':>4} {'FP':>5} {'FN':>4} {'recall':>7} "
          f"{'prec':>6} {'FP/h':>7} {'onset':>7} {'offset':>7} {'IoU':>5}")
    for r in rows:
        f = lambda v, s="{:.1f}": s.format(v) if v is not None else "  -"
        print(f"  {r['thr']:>5} {r['n_det']:>6} {r['tp']:>4} {r['fp']:>5} {r['fn']:>4} "
              f"{r['recall']:>7.3f} {r['precision']:>6.3f} {r['fp_per_hour']:>7.1f} "
              f"{f(r['onset_mae']):>7} {f(r['offset_mae']):>7} {f(r['mean_iou'],'{:.2f}'):>5}")

    (out / "results.json").write_text(json.dumps(
        dict(session=args.session, subject=args.subject, ckpt=args.ckpt,
             hours=hours, n_annotated=len(gt), sweep=rows), indent=2))
    print(f"\nwrote {out}/results.json  and  {out}/trace.npz")


if __name__ == "__main__":
    main()
