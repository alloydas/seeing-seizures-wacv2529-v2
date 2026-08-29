
import os, re, sys, glob, subprocess, csv, warnings, argparse
from datetime import datetime, timedelta
from pathlib import Path

warnings.filterwarnings("ignore", category=RuntimeWarning)

# ─────────────────────────────────────────────────────────────
#  CLI ARGUMENTS  (all former config block is now command-line)
# ─────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="Seizure Clip Extractor — cuts video + EEG clips for every seizure annotation.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Batch: process all session subfolders under a root directory
  python cut_seizure_clips.py --batch /path/to/archive

  # Single session: explicit paths
  python cut_seizure_clips.py --session /path/to/repo/12-11-2023

  # Single session with custom output and buffers
  python cut_seizure_clips.py --session /path/to/repo/12-11-2023 \\
      --output /path/to/repo/clips --pre 5 --post 20 --mode reencode

  # Batch, only Stage 4 seizures
  python cut_seizure_clips.py --batch /path/to/archive \\
      --filter "Stage 4"
        """
    )

    src = p.add_mutually_exclusive_group(required=True)
    src.add_argument("--batch",   metavar="DIR",
                     help="Root folder whose subdirectories are session folders")
    src.add_argument("--session", metavar="DIR",
                     help="Single session folder (contains .mp4, .edf, .xlsx)")

    p.add_argument("--output",  metavar="DIR",  default=None,
                   help="Output folder (default: <session>/seizure_clips)")
    p.add_argument("--pre",     metavar="S",    type=float, default=10,
                   help="Seconds of pre-seizure buffer (default: 10)")
    p.add_argument("--post",    metavar="S",    type=float, default=10,
                   help="Seconds of post-seizure buffer (default: 10)")
    p.add_argument("--mode",    choices=["copy", "reencode"], default="copy",
                   help="Video cut mode: copy=fast, reencode=frame-accurate (default: copy)")
    p.add_argument("--filter",  metavar="LABEL", nargs="*", default=[],
                   help='Only process seizures with these labels e.g. --filter "Stage 3" "Stage 4"')
    p.add_argument("--year",    metavar="YYYY",  type=int, default=2023,
                   help="Fallback year for EDF filenames without a year (default: 2023)")

    return p.parse_args()


args = parse_args()

# Make args available as module-level names so the rest of the code
# (parse_edf_files, parse_annotations, process_session, etc.) can read them.
PRE_BUFFER_S   = args.pre
POST_BUFFER_S  = args.post
CUT_MODE       = args.mode
FILTER_LABELS  = args.filter or []
RECORDING_YEAR = args.year

# ─────────────────────────────────────────────────────────────


# ══════════════════════════════════════════════════════════════
#  DEPENDENCY CHECK
# ══════════════════════════════════════════════════════════════

def check_deps():
    missing = []
    for pkg, pip_name in [("openpyxl",       "openpyxl"),
                           ("imageio_ffmpeg", "imageio-ffmpeg"),
                           ("cv2",            "opencv-python"),
                           ("mne",            "mne"),
                           ("edfio",          "edfio")]:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pip_name)
    if missing:
        print(f"\n[ERROR] Missing packages: {', '.join(missing)}")
        print(f"        Fix:  pip install {' '.join(missing)}")
        sys.exit(1)
    import imageio_ffmpeg, mne
    exe = imageio_ffmpeg.get_ffmpeg_exe()
    print(f"  Bundled ffmpeg : {exe}")
    print(f"  MNE version    : {mne.__version__}")
    return exe


# ══════════════════════════════════════════════════════════════
#  VIDEO FILE PARSING
# ══════════════════════════════════════════════════════════════

def _xml_start_time(fpath):
    """
    Look for a sidecar XML file (same base name, .XML or .xml extension)
    and read DSI_utc_start_time.  Returns a local datetime or None.

    DSI_local_start_time format: ID:0INDEXED_MONTH:DAY:HH:MM:SS
    The last three colon-separated values are the exact local HH:MM:SS.
    DSI_utc_start_time (Unix timestamp) converted with datetime.fromtimestamp()
    gives the identical local time plus the correct date, so we use that.
    """
    try:
        from xml.etree import ElementTree as ET
        base = os.path.splitext(fpath)[0]
        xml_path = None
        for ext in (".XML", ".xml"):
            candidate = base + ext
            if os.path.exists(candidate):
                xml_path = candidate
                break
        if xml_path is None:
            return None
        root = ET.parse(xml_path).getroot()
        utc_node = root.find("DSI_utc_start_time")
        if utc_node is not None and utc_node.text:
            return datetime.fromtimestamp(int(utc_node.text.strip()))
    except Exception:
        pass
    return None


def parse_video_files(folder):
    """
    Finds .mp4 files with a 14-digit (YYYYMMDDHHMMSS) or
    12-digit (YYYYMMDDHHMM) timestamp anywhere in the filename.
    If a sidecar .XML file exists, its DSI_utc_start_time overrides
    the filename-derived start time (more accurate local timestamp).
    """
    p14 = re.compile(r'(\d{14})')
    p12 = re.compile(r'(\d{12})')
    videos, unmatched = [], []

    all_files = sorted(set(
        glob.glob(os.path.join(folder, "*.mp4")) +
        glob.glob(os.path.join(folder, "*", "*.mp4"))
    ))

    for fpath in all_files:
        fname    = os.path.basename(fpath)
        start_dt = None

        # Prefer XML sidecar timestamp (exact local time)
        xml_dt = _xml_start_time(fpath)
        if xml_dt is not None:
            start_dt = xml_dt
            print(f"    [XML] {fname}: start={start_dt} (from sidecar XML)")

        # Fall back to filename-embedded timestamp
        if start_dt is None:
            for pat, fmt in [(p14, "%Y%m%d%H%M%S"), (p12, "%Y%m%d%H%M")]:
                m = pat.search(fname)
                if m:
                    try:
                        start_dt = datetime.strptime(m.group(1), fmt)
                        break
                    except ValueError:
                        pass

        if start_dt:
            videos.append((start_dt, fpath, fname))
        else:
            unmatched.append(fname)

    if unmatched:
        print(f"\n  [!] {len(unmatched)} .mp4 file(s) had no parseable timestamp:")
        for u in unmatched[:5]:
            print(f"      {u}")
        if len(unmatched) > 5:
            print(f"      ... and {len(unmatched)-5} more")

    return sorted(videos, key=lambda x: x[0])


def get_video_info(fpath, ffmpeg_exe):
    import cv2
    cap    = cv2.VideoCapture(fpath)
    fps    = cap.get(cv2.CAP_PROP_FPS)
    frames = cap.get(cv2.CAP_PROP_FRAME_COUNT)
    width  = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    cap.release()
    dur_s = frames / fps if fps > 0 and frames > 0 else None

    if not dur_s:
        try:
            r = subprocess.run([ffmpeg_exe, "-i", fpath],
                               capture_output=True, text=True, timeout=15)
            m = re.search(r'Duration: (\d+):(\d+):([\d.]+)', r.stderr)
            if m:
                h, mn, s = int(m.group(1)), int(m.group(2)), float(m.group(3))
                dur_s = h*3600 + mn*60 + s
        except Exception:
            pass

    return dur_s or 0, round(fps, 3), width, height


def build_video_intervals(videos, ffmpeg_exe):
    intervals = []
    for i, (start_dt, fpath, fname) in enumerate(videos):
        dur_s, fps, width, height = get_video_info(fpath, ffmpeg_exe)
        if dur_s and dur_s > 0:
            end_dt = start_dt + timedelta(seconds=dur_s)
        elif i + 1 < len(videos):
            end_dt = videos[i + 1][0]
            dur_s  = (end_dt - start_dt).total_seconds()
        else:
            end_dt = start_dt + timedelta(hours=4)
            dur_s  = 14400
        intervals.append({
            "start": start_dt, "end": end_dt, "dur_s": dur_s,
            "fps": fps, "width": width, "height": height,
            "fpath": fpath, "fname": fname
        })
    return intervals


# ══════════════════════════════════════════════════════════════
#  EDF FILE PARSING
# ══════════════════════════════════════════════════════════════

def _parse_edf_start_from_filename(fname):
    """
    Fallback date parser when the EDF header has no meas_date.

    Handles these formats (all real filenames observed):
      RN140-22 (12-13 to 12-15).edf   → 2022-12-13 00:00:00
      RN140-22 (12-15 to 12-16).edf   → 2022-12-15 00:00:00
      RN103-22 (10-28).edf             → 2022-10-28 00:00:00
      RN103_20221028.edf               → 2022-10-28 00:00:00
    """

    # ── Pattern 1: (MM-DD to MM-DD)  e.g. (12-13 to 12-15) ──────
    # Use the START date of the range
    m = re.search(r'\((\d{1,2})-(\d{1,2})\s+to\s+\d{1,2}-\d{1,2}\)', fname)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        try:
            return datetime(RECORDING_YEAR, month, day, 0, 0, 0)
        except ValueError:
            pass

    # ── Pattern 2: single (MM-DD)  e.g. (10-28) ─────────────────
    m = re.search(r'\((\d{1,2})-(\d{1,2})\)', fname)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        try:
            return datetime(RECORDING_YEAR, month, day, 0, 0, 0)
        except ValueError:
            pass

    # ── Pattern 3: full date YYYYMMDD anywhere in name ───────────
    m = re.search(r'(\d{4})(\d{2})(\d{2})', fname)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    return None


def parse_edf_files(folder):
    """
    Scans for .edf files, reads start time from the EDF header (meas_date).
    Falls back to parsing the filename if the header has no date.
    """
    import mne
    edf_records = []
    all_files   = sorted(set(
        glob.glob(os.path.join(folder, "*.edf")) +
        glob.glob(os.path.join(folder, "*.EDF")) +
        glob.glob(os.path.join(folder, "*", "*.edf"))
    ))

    if not all_files:
        print("  [!] No .edf files found. EEG segments will be skipped.")
        return []

    print(f"  Found {len(all_files)} EDF file(s):")
    for fpath in all_files:
        fname = os.path.basename(fpath)
        try:
            raw      = mne.io.read_raw_edf(fpath, preload=False, verbose=False)
            start_dt = raw.info["meas_date"]

            if start_dt is not None:
                # MNE returns timezone-aware datetime — strip timezone
                start_dt = start_dt.replace(tzinfo=None)
                source   = "header"
            else:
                # Fallback: parse from filename
                start_dt = _parse_edf_start_from_filename(fname)
                source   = "filename"
                if start_dt is None:
                    print(f"    [ERROR] {fname} — cannot determine start time")
                    raw.close()
                    continue

            dur_s  = raw.n_times / raw.info["sfreq"]
            end_dt = start_dt + timedelta(seconds=dur_s)
            n_ch   = len(raw.ch_names)
            sfreq  = raw.info["sfreq"]
            raw.close()

            edf_records.append({
                "start": start_dt, "end": end_dt, "dur_s": dur_s,
                "n_ch": n_ch, "sfreq": sfreq,
                "fpath": fpath, "fname": fname, "source": source
            })
            print(f"    {fname:<42}  {start_dt}  →  {end_dt}"
                  f"  ({dur_s/3600:.1f}h  {n_ch}ch  {sfreq}Hz  [{source}])")

        except Exception as e:
            print(f"    [ERROR] {fname}: {e}")

    edf_records.sort(key=lambda x: x["start"])
    return edf_records


def find_edf(sz_start, sz_end, edf_records):
    """Find which EDF file covers the seizure start timestamp."""
    for e in edf_records:
        if e["start"] <= sz_start < e["end"]:
            return e
    return None


# ══════════════════════════════════════════════════════════════
#  ANNOTATION PARSING
# ══════════════════════════════════════════════════════════════

def parse_annotations(excel_path):
    import openpyxl
    wb      = openpyxl.load_workbook(excel_path)
    ws      = wb.active
    headers = [str(c.value).strip().lower() if c.value else ""
               for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"  Columns: {headers}")

    def col(*names):
        for n in names:
            for i, h in enumerate(headers):
                if n in h:
                    return i
        return None

    idx_label = col("label", "stage", "type")
    idx_dur   = col("dur", "length")
    idx_start = col("start")
    idx_end   = col("end")
    idx_spike = col("spike")

    if idx_start is None:
        sys.exit(f"[ERROR] No 'Start' column found. Headers: {headers}")

    seizures = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        start = row[idx_start] if idx_start is not None else None
        if start is None:
            continue
        # Skip unit-label rows (e.g. a second header row with "[s]" under duration)
        if not isinstance(start, (int, float)) and not hasattr(start, 'hour'):
            try:
                datetime.strptime(str(start).strip(), "%H:%M:%S")
            except Exception:
                continue
        label    = row[idx_label] if idx_label is not None else None
        label    = str(label).strip() if label is not None else "Unknown"
        dur_raw  = row[idx_dur] if idx_dur is not None else None
        try:
            duration = float(dur_raw) if dur_raw is not None else 30.0
        except (ValueError, TypeError):
            duration = 30.0
        end      = row[idx_end]   if idx_end   is not None else None
        spikes   = row[idx_spike] if idx_spike is not None else 0
        if FILTER_LABELS and label not in FILTER_LABELS:
            continue
        seizures.append({
            "idx": i+1, "label": label, "duration": duration,
            "start": start, "end": end, "spikes": spikes
        })
    return seizures


# ══════════════════════════════════════════════════════════════
#  CUTTING FUNCTIONS
# ══════════════════════════════════════════════════════════════

def seconds_to_hms(s):
    s = max(0, int(s))
    return f"{s//3600:02d}:{(s%3600)//60:02d}:{s%60:02d}"


def find_video(sz_start, intervals):
    for v in intervals:
        if v["start"] <= sz_start < v["end"]:
            return v
    return None


def cut_video_clip(ffmpeg_exe, video_path, seek_s, clip_dur_s, out_path, mode="copy"):
    seek_s = max(0.0, float(seek_s))
    if mode == "copy":
        cmd = [ffmpeg_exe, "-y",
               "-ss", str(seek_s), "-i", video_path,
               "-t", str(clip_dur_s),
               "-c", "copy", "-avoid_negative_ts", "make_zero",
               out_path]
    else:
        cmd = [ffmpeg_exe, "-y",
               "-i", video_path,
               "-ss", str(seek_s), "-t", str(clip_dur_s),
               "-c:v", "libx264", "-crf", "18",
               "-c:a", "aac", "-preset", "fast",
               out_path]
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
    return r.returncode == 0, r.stderr


def cut_eeg_segment(edf_record, clip_start_dt, clip_end_dt, out_path):
    """
    Loads the EDF, crops to [clip_start_dt, clip_end_dt], exports as .edf.
    """
    import mne
    try:
        # preload=False + crop + load_data reads ONLY the clip window from disk.
        # preload=True here re-read the entire multi-day EDF (up to ~1.5 GB) for
        # every single clip, so a 231-seizure session did ~350 GB of reads and
        # blew past the cut timeout. Same output, ~100x less I/O on big EDFs.
        raw      = mne.io.read_raw_edf(edf_record["fpath"], preload=False, verbose=False)
        edf_start = edf_record["start"]

        t_start_s = max(0.0, (clip_start_dt - edf_start).total_seconds())
        t_end_s   = min(edf_record["dur_s"],
                        (clip_end_dt   - edf_start).total_seconds())

        if t_end_s <= t_start_s:
            raw.close()
            return False, f"Bad crop window [{t_start_s:.1f}s → {t_end_s:.1f}s]"

        raw.crop(tmin=t_start_s, tmax=t_end_s)
        raw.load_data(verbose=False)   # loads only the cropped window

        # EDF standard: channel names must be ≤16 characters
        rename_map = {}
        used_names = set(raw.ch_names)
        for ch in raw.ch_names:
            if len(ch) > 16:
                base = ch[:14]
                suffix = 0
                candidate = ch[:16]
                while candidate in used_names and candidate != ch:
                    candidate = f"{base}_{suffix}"
                    suffix += 1
                rename_map[ch] = candidate
                used_names.discard(ch)
                used_names.add(candidate)
        if rename_map:
            raw.rename_channels(rename_map)

        # mne.export.export_raw writes a SHARED physical range across channels.
        # With SignalStr (~31) and Temp (~37) in the same file, one 16-bit step
        # became larger than the entire ECG/EEG amplitude and the biopotential
        # collapsed to 2-4 distinct values -- 69% of all cut clips (2026-08-12
        # audit). write_edf_clip gives each channel its own range; verified to
        # reproduce the source at correlation 1.000000.
        from edf_clip_writer import write_edf_clip
        write_edf_clip(raw, out_path)
        raw.close()

        dur    = t_end_s - t_start_s
        size   = os.path.getsize(out_path) / (1024*1024)
        return True, (f"{edf_record['n_ch']}ch  "
                      f"{edf_record['sfreq']}Hz  "
                      f"{dur:.1f}s  ~{size:.1f}MB")
    except Exception as e:
        return False, str(e)


# ══════════════════════════════════════════════════════════════
#  CORE SESSION PROCESSOR
# ══════════════════════════════════════════════════════════════

def process_session(session_folder, ffmpeg_exe, excel_file=None, output_folder=None):
    """
    Process one session folder: find videos, EDF, and annotations,
    then cut clips for every seizure.  Returns summary counts dict.
    """
    session_folder = str(session_folder)
    output_folder  = output_folder or os.path.join(session_folder, "seizure_clips")

    # Auto-detect excel if not provided
    if excel_file is None:
        xlsx_files = (glob.glob(os.path.join(session_folder, "*.xlsx")) +
                      glob.glob(os.path.join(session_folder, "*.XLSX")))
        if not xlsx_files:
            print(f"  [SKIP] No .xlsx found in {session_folder}")
            return None
        excel_file = xlsx_files[0]
        if len(xlsx_files) > 1:
            print(f"  [!] Multiple .xlsx found, using: {os.path.basename(excel_file)}")

    # 1. Videos
    print(f"\n  [1] Videos: {session_folder}")
    videos    = parse_video_files(session_folder)
    vid_ivals = build_video_intervals(videos, ffmpeg_exe)
    print(f"      {len(vid_ivals)} video file(s)")
    for v in vid_ivals:
        res = f"{v['width']}x{v['height']}" if v['width'] else "?"
        print(f"      {v['fname']:<42} {str(v['start']):<22} "
              f"{seconds_to_hms(v['dur_s'])}  {v['fps']}fps  {res}")

    # 2. EDF
    print(f"\n  [2] EDF: {session_folder}")
    edf_records = parse_edf_files(session_folder)

    # 3. Annotations
    print(f"\n  [3] Annotations: {os.path.basename(excel_file)}")
    seizures = parse_annotations(excel_file)
    print(f"      {len(seizures)} seizure(s) loaded")

    if not seizures:
        print(f"  [SKIP] No seizures in {os.path.basename(excel_file)}")
        return {"ok_both": 0, "ok_vid_only": 0, "ok_eeg_only": 0,
                "skipped": 0, "failed": 0, "total": 0}

    # 4. Output folder
    os.makedirs(output_folder, exist_ok=True)
    print(f"\n  [4] Output → {output_folder}")

    ok_both = ok_vid_only = ok_eeg_only = skipped = failed = 0
    log_rows = []

    for sz in seizures:
        idx      = sz["idx"]
        label    = sz["label"]
        sz_start = sz["start"]
        sz_end   = sz["end"] if sz["end"] else sz_start + timedelta(seconds=sz["duration"])
        sz_dur   = sz["duration"]

        clip_start = sz_start - timedelta(seconds=PRE_BUFFER_S)
        clip_end   = sz_end   + timedelta(seconds=POST_BUFFER_S)
        clip_dur_s = (clip_end - clip_start).total_seconds()

        video = find_video(sz_start, vid_ivals)
        edf   = find_edf(sz_start, sz_end, edf_records)

        if video is None and edf is None:
            gap_str = "—"
            if vid_ivals:
                nearest_v = min(vid_ivals,
                    key=lambda v: abs((v["start"] - sz_start).total_seconds()))
                gap_str = f"nearest video gap {(sz_start - nearest_v['end']).total_seconds():.0f}s"
            print(f"  [SKIP]  #{idx:02d}  {sz_start}  — no video or EEG coverage")
            log_rows.append([idx, label, str(sz_start), "—", "—",
                             "NO_COVERAGE", "—", gap_str])
            skipped += 1
            continue

        safe_label  = re.sub(r'[^\w]', '_', label)
        folder_name = (f"seizure_{idx:02d}_{safe_label}"
                       f"_{sz_start.strftime('%Y%m%d_%H%M%S')}")
        clip_folder = os.path.join(output_folder, folder_name)
        os.makedirs(clip_folder, exist_ok=True)

        print(f"  ── #{idx:02d}  {sz_start}  ({sz_dur:.1f}s  {label})")

        vid_status = eeg_status = "—"

        # ── Cut video ──────────────────────────────────────────
        if video:
            raw_seek   = (sz_start - video["start"]).total_seconds()
            seek_s     = max(0.0, raw_seek - PRE_BUFFER_S)
            v_clip_dur = min(clip_dur_s, video["dur_s"] - seek_s)
            vid_out    = os.path.join(clip_folder, "video.mp4")

            if os.path.exists(vid_out):
                print(f"     VIDEO : [already exists]")
                vid_status = "EXISTS"
            else:
                print(f"     VIDEO : seek={seconds_to_hms(raw_seek)}  "
                      f"dur={v_clip_dur:.0f}s  ", end="", flush=True)
                ok, err = cut_video_clip(ffmpeg_exe, video["fpath"],
                                          seek_s, v_clip_dur, vid_out, CUT_MODE)
                if ok and os.path.exists(vid_out):
                    mb  = os.path.getsize(vid_out) / (1024*1024)
                    fps = str(video["fps"]) if video["fps"] else "?"
                    print(f"OK  ({mb:.1f}MB  {fps}fps)")
                    vid_status = f"OK {mb:.1f}MB"
                else:
                    errmsg = (err.strip().splitlines() or [""])[-1][:60]
                    print(f"FAILED — {errmsg}")
                    vid_status = "FAILED"
        else:
            print(f"     VIDEO : no matching video file")

        # ── Cut EEG ────────────────────────────────────────────
        if edf:
            eeg_out = os.path.join(clip_folder, "eeg.edf")
            if os.path.exists(eeg_out):
                print(f"     EEG   : [already exists]")
                eeg_status = "EXISTS"
            else:
                print(f"     EEG   : {edf['fname']}  "
                      f"(start from {edf['source']})  ", end="", flush=True)
                ok, msg = cut_eeg_segment(edf, clip_start, clip_end, eeg_out)
                if ok:
                    print(f"OK  ({msg})")
                    eeg_status = f"OK {msg}"
                else:
                    print(f"FAILED — {msg}")
                    eeg_status = f"FAILED: {msg}"
        else:
            print(f"     EEG   : no matching EDF file")

        # ── Write info.txt ─────────────────────────────────────
        with open(os.path.join(clip_folder, "info.txt"), "w") as f:
            f.write(f"Seizure index  : {idx}\n")
            f.write(f"Label          : {label}\n")
            f.write(f"Seizure start  : {sz_start}\n")
            f.write(f"Seizure end    : {sz_end}\n")
            f.write(f"Duration (s)   : {sz_dur:.2f}\n")
            f.write(f"Spikes         : {sz['spikes']}\n")
            f.write(f"Clip start     : {clip_start}\n")
            f.write(f"Clip end       : {clip_end}\n")
            f.write(f"Pre-buffer (s) : {PRE_BUFFER_S}\n")
            f.write(f"Post-buffer(s) : {POST_BUFFER_S}\n")
            f.write(f"Video file     : {video['fname'] if video else 'N/A'}\n")
            f.write(f"EDF file       : {edf['fname']  if edf   else 'N/A'}\n")
            if edf:
                f.write(f"EEG channels   : {edf['n_ch']}\n")
                f.write(f"EEG sfreq (Hz) : {edf['sfreq']}\n")
                f.write(f"EDF date src   : {edf['source']}\n")
            if video:
                f.write(f"Video FPS      : {video['fps']}\n")
                f.write(f"Video res      : {video['width']}x{video['height']}\n")

        v_ok = "OK" in vid_status
        e_ok = "OK" in eeg_status
        if v_ok and e_ok:   ok_both += 1
        elif v_ok:          ok_vid_only += 1
        elif e_ok:          ok_eeg_only += 1
        else:               failed += 1

        log_rows.append([idx, label, str(sz_start),
                         seconds_to_hms(int(
                             (sz_start - (video["start"] if video else sz_start)
                              ).total_seconds())),
                         f"{clip_dur_s:.1f}s",
                         vid_status, eeg_status, folder_name])

    # Log CSV
    log_path = os.path.join(output_folder, "_clip_log.csv")
    with open(log_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["#", "Label", "Seizure Start", "Video Seek",
                    "Clip Duration", "Video Status", "EEG Status", "Folder"])
        w.writerows(log_rows)

    print(f"  Log → {log_path}")
    return {"ok_both": ok_both, "ok_vid_only": ok_vid_only,
            "ok_eeg_only": ok_eeg_only, "skipped": skipped,
            "failed": failed, "total": len(seizures)}


# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════

def main():
    print("\n" + "="*72)
    print("  SEIZURE CLIP EXTRACTOR  v3  —  Video + EEG")
    print("="*72)
    print(f"  Pre={PRE_BUFFER_S}s | Post={POST_BUFFER_S}s | Mode={CUT_MODE}"
          + (f" | Filter={FILTER_LABELS}" if FILTER_LABELS else ""))

    ffmpeg_exe = check_deps()

    if args.batch:
        # ── Batch mode: process every subdirectory ─────────────
        base = Path(args.batch)
        if not base.is_dir():
            sys.exit(f"[ERROR] --batch path does not exist: {base}")
        sessions = sorted([d for d in base.iterdir() if d.is_dir()])
        print(f"\n  Batch root : {base}")
        print(f"  Sessions   : {len(sessions)}\n")

        totals = {"ok_both": 0, "ok_vid_only": 0, "ok_eeg_only": 0,
                  "skipped": 0, "failed": 0, "total": 0}

        for i, session in enumerate(sessions, 1):
            print("\n" + "─"*72)
            print(f"  SESSION {i}/{len(sessions)}: {session.name}")
            print("─"*72)
            out = Path(args.output) / session.name if args.output else None
            result = process_session(str(session), ffmpeg_exe,
                                     output_folder=str(out) if out else None)
            if result:
                for k in totals:
                    totals[k] += result[k]

        print("\n" + "="*72)
        print("  BATCH COMPLETE — OVERALL SUMMARY")
        print("="*72)
        print(f"  Sessions processed : {len(sessions)}")
        print(f"  Both video + EEG   : {totals['ok_both']}")
        print(f"  Video only         : {totals['ok_vid_only']}")
        print(f"  EEG only           : {totals['ok_eeg_only']}")
        print(f"  Skipped            : {totals['skipped']}")
        print(f"  Failed             : {totals['failed']}")
        print(f"  Total seizures     : {totals['total']}\n")

    else:
        # ── Single-session mode ────────────────────────────────
        session = Path(args.session)
        if not session.is_dir():
            sys.exit(f"[ERROR] --session path does not exist: {session}")
        out = args.output or str(session / "seizure_clips")

        result = process_session(str(session), ffmpeg_exe, output_folder=out)
        r = result or {}
        print("\n" + "="*72)
        print("  SUMMARY")
        print("="*72)
        print(f"  Both video + EEG : {r.get('ok_both', 0)}")
        print(f"  Video only       : {r.get('ok_vid_only', 0)}")
        print(f"  EEG only         : {r.get('ok_eeg_only', 0)}")
        print(f"  Skipped          : {r.get('skipped', 0)}")
        print(f"  Failed           : {r.get('failed', 0)}")
        print(f"\n  Output → {os.path.abspath(out)}\n")


if __name__ == "__main__":
    main()
