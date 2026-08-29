"""
Non-Seizure Clip Sampler
========================
For every seizure annotation in a session, samples a random clip of the
same total duration (pre + seizure + post) from a time window that does NOT
overlap any annotated seizure (± a configurable safety margin).

Output structure mirrors cut_seizure_clips.py:
    non_seizure_clips/
        clip_01_vs_seizure_02_Stage_3_20231012_162605/
            video.mp4
            eeg.edf
            info.txt
        _non_seizure_clip_log.csv

Usage examples:
  # Single session
  python cut_non_seizure_clips.py --session /path/to/archive

  # Batch (every subdirectory)
  python cut_non_seizure_clips.py --batch /path/to/archive

  # Custom buffers, reencode, filter to Stage 3 only
  python cut_non_seizure_clips.py --batch /path/to/archive \\
      --pre 10 --post 10 --safety 30 --mode reencode --filter "Stage 3"

  # Fixed random seed for reproducibility
  python cut_non_seizure_clips.py --batch /path/to/archive --seed 42
"""

import os, re, sys, glob, subprocess, csv, warnings, argparse, random
from datetime import datetime, timedelta
from pathlib import Path

warnings.filterwarnings("ignore", category=RuntimeWarning)


# ─────────────────────────────────────────────────────────────
#  CLI ARGUMENTS
# ─────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="Non-Seizure Clip Sampler — cuts random background clips matching each seizure's duration.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    src = p.add_mutually_exclusive_group(required=True)
    src.add_argument("--batch",   metavar="DIR",
                     help="Root folder whose subdirectories are session folders")
    src.add_argument("--session", metavar="DIR",
                     help="Single session folder (contains .mp4, .edf, .xlsx)")

    p.add_argument("--output",  metavar="DIR",  default=None,
                   help="Output folder (default: <session>/non_seizure_clips)")
    p.add_argument("--pre",     metavar="S",    type=float, default=10,
                   help="Seconds of pre-seizure buffer used in seizure clips (default: 10)")
    p.add_argument("--post",    metavar="S",    type=float, default=10,
                   help="Seconds of post-seizure buffer used in seizure clips (default: 10)")
    p.add_argument("--safety",  metavar="S",    type=float, default=30,
                   help="Extra seconds to keep clear on each side of every seizure (default: 30)")
    p.add_argument("--mode",    choices=["copy", "reencode"], default="copy",
                   help="Video cut mode: copy=fast, reencode=frame-accurate (default: copy)")
    p.add_argument("--filter",  metavar="LABEL", nargs="*", default=[],
                   help='Only create non-seizure clips to match seizures with these labels')
    p.add_argument("--year",    metavar="YYYY",  type=int, default=2023,
                   help="Fallback year for EDF filenames without a year (default: 2023)")
    p.add_argument("--seed",    metavar="N",     type=int, default=None,
                   help="Random seed for reproducible sampling (default: random)")
    return p.parse_args()


args = parse_args()

PRE_BUFFER_S   = args.pre
POST_BUFFER_S  = args.post
SAFETY_S       = args.safety
CUT_MODE       = args.mode
FILTER_LABELS  = args.filter or []
RECORDING_YEAR = args.year

if args.seed is not None:
    random.seed(args.seed)


# ─────────────────────────────────────────────────────────────
#  DEPENDENCY CHECK
# ─────────────────────────────────────────────────────────────

def check_deps():
    missing = []
    for pkg, pip_name in [("openpyxl",       "openpyxl"),
                           ("imageio_ffmpeg", "imageio-ffmpeg"),
                           ("cv2",            "opencv-python"),
                           ("mne",            "mne"),
                           ("edfio",          "edfio")]:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pip_name)
    if missing:
        sys.exit(f"\n[ERROR] Missing packages: {', '.join(missing)}\n"
                 f"        Fix:  pip install {' '.join(missing)}")
    import imageio_ffmpeg, mne
    exe = imageio_ffmpeg.get_ffmpeg_exe()
    print(f"  Bundled ffmpeg : {exe}")
    print(f"  MNE version    : {mne.__version__}")
    return exe


# ─────────────────────────────────────────────────────────────
#  VIDEO FILE PARSING
# ─────────────────────────────────────────────────────────────

def _xml_start_time(fpath):
    """Read DSI_utc_start_time from sidecar .XML file if present."""
    try:
        from xml.etree import ElementTree as ET
        base = os.path.splitext(fpath)[0]
        for ext in (".XML", ".xml"):
            candidate = base + ext
            if os.path.exists(candidate):
                root = ET.parse(candidate).getroot()
                node = root.find("DSI_utc_start_time")
                if node is not None and node.text:
                    return datetime.fromtimestamp(int(node.text.strip()))
    except Exception:
        pass
    return None


def parse_video_files(folder):
    p14 = re.compile(r'(\d{14})')
    p12 = re.compile(r'(\d{12})')
    videos, unmatched = [], []

    all_files = sorted(set(
        glob.glob(os.path.join(folder, "*.mp4")) +
        glob.glob(os.path.join(folder, "*", "*.mp4"))
    ))

    for fpath in all_files:
        fname    = os.path.basename(fpath)
        start_dt = _xml_start_time(fpath)   # prefer XML sidecar

        if start_dt is None:
            for pat, fmt in [(p14, "%Y%m%d%H%M%S"), (p12, "%Y%m%d%H%M")]:
                m = pat.search(fname)
                if m:
                    try:
                        start_dt = datetime.strptime(m.group(1), fmt)
                        break
                    except ValueError:
                        pass

        if start_dt:
            videos.append((start_dt, fpath, fname))
        else:
            unmatched.append(fname)

    if unmatched:
        print(f"  [!] {len(unmatched)} .mp4 file(s) had no parseable timestamp (skipped)")

    return sorted(videos, key=lambda x: x[0])


def get_video_info(fpath, ffmpeg_exe):
    import cv2
    cap    = cv2.VideoCapture(fpath)
    fps    = cap.get(cv2.CAP_PROP_FPS)
    frames = cap.get(cv2.CAP_PROP_FRAME_COUNT)
    width  = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    cap.release()
    dur_s = frames / fps if fps > 0 and frames > 0 else None

    if not dur_s:
        try:
            r = subprocess.run([ffmpeg_exe, "-i", fpath],
                               capture_output=True, text=True, timeout=15)
            m = re.search(r'Duration: (\d+):(\d+):([\d.]+)', r.stderr)
            if m:
                h, mn, s = int(m.group(1)), int(m.group(2)), float(m.group(3))
                dur_s = h * 3600 + mn * 60 + s
        except Exception:
            pass
    return dur_s or 0, round(fps, 3), width, height


def build_video_intervals(videos, ffmpeg_exe):
    intervals = []
    for i, (start_dt, fpath, fname) in enumerate(videos):
        dur_s, fps, width, height = get_video_info(fpath, ffmpeg_exe)
        if dur_s > 0:
            end_dt = start_dt + timedelta(seconds=dur_s)
        elif i + 1 < len(videos):
            end_dt = videos[i + 1][0]
            dur_s  = (end_dt - start_dt).total_seconds()
        else:
            end_dt = start_dt + timedelta(hours=4)
            dur_s  = 14400
        res = f"{width}x{height}" if width else "?"
        print(f"    {fname:<45}  {str(start_dt):<22}  {_hms(dur_s)}"
              f"  {fps}fps  {res}")
        intervals.append({
            "start": start_dt, "end": end_dt, "dur_s": dur_s,
            "fps": fps, "width": width, "height": height,
            "fpath": fpath, "fname": fname,
        })
    return intervals


# ─────────────────────────────────────────────────────────────
#  EDF FILE PARSING
# ─────────────────────────────────────────────────────────────

def _parse_edf_start_from_filename(fname):
    m = re.search(r'\((\d{1,2})-(\d{1,2})\s+to\s+\d{1,2}-\d{1,2}\)', fname)
    if m:
        try: return datetime(RECORDING_YEAR, int(m.group(1)), int(m.group(2)))
        except ValueError: pass
    m = re.search(r'\((\d{1,2})-(\d{1,2})\)', fname)
    if m:
        try: return datetime(RECORDING_YEAR, int(m.group(1)), int(m.group(2)))
        except ValueError: pass
    m = re.search(r'(\d{4})(\d{2})(\d{2})', fname)
    if m:
        try: return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError: pass
    return None


def parse_edf_files(folder):
    import mne
    edf_records = []
    all_files = sorted(set(
        glob.glob(os.path.join(folder, "*.edf")) +
        glob.glob(os.path.join(folder, "*.EDF")) +
        glob.glob(os.path.join(folder, "*", "*.edf"))
    ))
    if not all_files:
        print("  [!] No .edf files found — EEG segments will be skipped.")
        return []

    print(f"  Found {len(all_files)} EDF file(s):")
    for fpath in all_files:
        fname = os.path.basename(fpath)
        try:
            raw      = mne.io.read_raw_edf(fpath, preload=False, verbose=False)
            start_dt = raw.info["meas_date"]
            if start_dt is not None:
                start_dt = start_dt.replace(tzinfo=None)
                source   = "header"
            else:
                start_dt = _parse_edf_start_from_filename(fname)
                source   = "filename"
                if start_dt is None:
                    print(f"    [ERROR] {fname} — cannot determine start time")
                    raw.close(); continue
            dur_s  = raw.n_times / raw.info["sfreq"]
            end_dt = start_dt + timedelta(seconds=dur_s)
            n_ch   = len(raw.ch_names)
            sfreq  = raw.info["sfreq"]
            raw.close()
            edf_records.append({
                "start": start_dt, "end": end_dt, "dur_s": dur_s,
                "n_ch": n_ch, "sfreq": sfreq,
                "fpath": fpath, "fname": fname, "source": source,
            })
            print(f"    {fname:<42}  {start_dt}  →  {end_dt}"
                  f"  ({dur_s/3600:.1f}h  {n_ch}ch  {sfreq}Hz  [{source}])")
        except Exception as e:
            print(f"    [ERROR] {fname}: {e}")

    edf_records.sort(key=lambda x: x["start"])
    return edf_records


def find_edf(sz_start, sz_end, edf_records):
    for e in edf_records:
        if e["start"] <= sz_start < e["end"]:
            return e
    return None


# ─────────────────────────────────────────────────────────────
#  ANNOTATION PARSING
# ─────────────────────────────────────────────────────────────

def parse_annotations(excel_path, filter_labels=None):
    """
    Reads seizure annotations.  Returns two lists:
      all_seizures  — every row (used to build the blocked-time map)
      target_seizures — rows matching filter_labels (we sample one clip for each)
    """
    import openpyxl
    wb      = openpyxl.load_workbook(excel_path)
    ws      = wb.active
    headers = [str(c.value).strip().lower() if c.value else ""
               for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"  Columns: {headers}")

    def col(*names):
        for n in names:
            for i, h in enumerate(headers):
                if n in h: return i
        return None

    idx_label = col("label", "stage", "type")
    idx_dur   = col("dur", "length")
    idx_start = col("start")
    idx_end   = col("end")
    idx_spike = col("spike")

    if idx_start is None:
        sys.exit(f"[ERROR] No 'Start' column found. Headers: {headers}")

    all_seizures, target_seizures = [], []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        start = row[idx_start] if idx_start is not None else None
        if start is None:
            continue
        # Skip unit-label rows (e.g. '[s]' under duration column)
        if not isinstance(start, (int, float)) and not hasattr(start, "hour"):
            try:
                datetime.strptime(str(start).strip(), "%H:%M:%S")
            except Exception:
                continue

        label    = row[idx_label] if idx_label is not None else None
        label    = str(label).strip() if label is not None else "Unknown"
        dur_raw  = row[idx_dur]  if idx_dur  is not None else None
        try:
            duration = float(dur_raw) if dur_raw is not None else 30.0
        except (ValueError, TypeError):
            duration = 30.0
        end    = row[idx_end]   if idx_end   is not None else None
        spikes = row[idx_spike] if idx_spike is not None else 0

        sz_end = end if end else start + timedelta(seconds=duration)

        record = {
            "idx": i + 1, "label": label, "duration": duration,
            "start": start, "end": sz_end, "spikes": spikes or 0,
        }
        all_seizures.append(record)
        if not filter_labels or label in filter_labels:
            target_seizures.append(record)

    return all_seizures, target_seizures


# ─────────────────────────────────────────────────────────────
#  FREE-WINDOW SAMPLING
# ─────────────────────────────────────────────────────────────

def build_blocked_intervals(all_seizures):
    """
    Returns sorted list of (block_start, block_end) datetimes representing
    all seizure regions expanded by PRE_BUFFER + SAFETY on each side.
    """
    margin = timedelta(seconds=PRE_BUFFER_S + SAFETY_S)
    blocked = []
    for sz in all_seizures:
        blocked.append((sz["start"] - margin,
                        sz["end"]   + timedelta(seconds=POST_BUFFER_S + SAFETY_S)))
    blocked.sort(key=lambda x: x[0])

    # Merge overlapping blocked intervals
    merged = []
    for b in blocked:
        if merged and b[0] <= merged[-1][1]:
            merged[-1] = (merged[-1][0], max(merged[-1][1], b[1]))
        else:
            merged.append(list(b))
    return [(datetime(*b[0].timetuple()[:6], b[0].microsecond),
             datetime(*b[1].timetuple()[:6], b[1].microsecond))
            for b in merged]


def find_free_windows(video, blocked_intervals, clip_dur_s):
    """
    For a single video interval, compute sub-windows that are completely
    free of seizures (and safety margin) and large enough to fit clip_dur_s.

    Returns list of (window_start_dt, window_end_dt) in absolute time.
    """
    vid_start = video["start"]
    vid_end   = video["end"]
    need      = timedelta(seconds=clip_dur_s)

    # Build the free segments within this video
    free_segments = []
    cursor = vid_start

    for b_start, b_end in blocked_intervals:
        # Clip blocked region to video bounds
        b_start = max(b_start, vid_start)
        b_end   = min(b_end,   vid_end)
        if b_start >= vid_end or b_end <= vid_start:
            continue
        if cursor < b_start:
            free_segments.append((cursor, b_start))
        cursor = max(cursor, b_end)

    if cursor < vid_end:
        free_segments.append((cursor, vid_end))

    # Keep only segments large enough for the clip
    usable = [(s, e) for s, e in free_segments if (e - s) >= need]
    return usable


def sample_random_window(free_windows, clip_dur_s):
    """
    Pick a random start time uniformly from all available free time,
    weighted by each window's usable length.  Returns absolute datetime.
    """
    need = timedelta(seconds=clip_dur_s)
    # Build weighted list: each free window contributes (end - need - start) of usable starts
    weights, starts, ends = [], [], []
    for s, e in free_windows:
        usable_end = e - need
        if usable_end > s:
            w = (usable_end - s).total_seconds()
            weights.append(w)
            starts.append(s)
            ends.append(usable_end)

    if not weights:
        return None

    total = sum(weights)
    r     = random.uniform(0, total)
    cum   = 0
    for w, s, e in zip(weights, starts, ends):
        cum += w
        if r <= cum:
            offset_s = random.uniform(0, w)
            return s + timedelta(seconds=offset_s)
    return starts[-1]   # fallback


# ─────────────────────────────────────────────────────────────
#  CUTTING FUNCTIONS
# ─────────────────────────────────────────────────────────────

def _hms(s):
    s = max(0, int(s))
    return f"{s//3600:02d}:{(s%3600)//60:02d}:{s%60:02d}"


def cut_video_clip(ffmpeg_exe, video_path, seek_s, clip_dur_s, out_path, mode="copy"):
    seek_s = max(0.0, float(seek_s))
    if mode == "copy":
        cmd = [ffmpeg_exe, "-y",
               "-ss", str(seek_s), "-i", video_path,
               "-t", str(clip_dur_s),
               "-c", "copy", "-avoid_negative_ts", "make_zero",
               out_path]
    else:
        cmd = [ffmpeg_exe, "-y",
               "-i", video_path,
               "-ss", str(seek_s), "-t", str(clip_dur_s),
               "-c:v", "libx264", "-crf", "18",
               "-c:a", "aac", "-preset", "fast",
               out_path]
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
    return r.returncode == 0, r.stderr


def cut_eeg_segment(edf_record, clip_start_dt, clip_end_dt, out_path):
    import mne
    try:
        # preload=False + crop + load_data reads ONLY the clip window (see the
        # matching fix in cut_seizure_clips.py); preload=True re-read the whole
        # multi-day EDF per clip and blew past the cut timeout on big EDFs.
        raw      = mne.io.read_raw_edf(edf_record["fpath"], preload=False, verbose=False)
        edf_start = edf_record["start"]
        t_start_s = max(0.0, (clip_start_dt - edf_start).total_seconds())
        t_end_s   = min(edf_record["dur_s"],
                        (clip_end_dt - edf_start).total_seconds())
        if t_end_s <= t_start_s:
            raw.close()
            return False, f"Bad crop window [{t_start_s:.1f}s → {t_end_s:.1f}s]"
        raw.crop(tmin=t_start_s, tmax=t_end_s)
        raw.load_data(verbose=False)   # loads only the cropped window

        # EDF standard: channel names must be ≤16 characters
        rename_map = {}
        used_names = set(raw.ch_names)
        for ch in raw.ch_names:
            if len(ch) > 16:
                base = ch[:14]
                suffix = 0
                candidate = ch[:16]
                while candidate in used_names and candidate != ch:
                    candidate = f"{base}_{suffix}"
                    suffix += 1
                rename_map[ch] = candidate
                used_names.discard(ch)
                used_names.add(candidate)
        if rename_map:
            raw.rename_channels(rename_map)

        # mne.export.export_raw writes a SHARED physical range across channels, which
        # crushed the biopotential to 2-4 distinct values in 69% of cut clips (see
        # edf_clip_writer.py). Per-channel ranges reproduce the source exactly.
        from edf_clip_writer import write_edf_clip
        write_edf_clip(raw, out_path)
        raw.close()
        dur  = t_end_s - t_start_s
        size = os.path.getsize(out_path) / (1024 * 1024)
        return True, (f"{edf_record['n_ch']}ch  "
                      f"{edf_record['sfreq']}Hz  "
                      f"{dur:.1f}s  ~{size:.1f}MB")
    except Exception as e:
        return False, str(e)


# ─────────────────────────────────────────────────────────────
#  CORE SESSION PROCESSOR
# ─────────────────────────────────────────────────────────────

def process_session(session_folder, ffmpeg_exe, output_folder=None):
    session_folder = str(session_folder)
    output_folder  = output_folder or os.path.join(session_folder, "non_seizure_clips")

    # Auto-detect xlsx
    xlsx_files = (glob.glob(os.path.join(session_folder, "*.xlsx")) +
                  glob.glob(os.path.join(session_folder, "*.XLSX")))
    if not xlsx_files:
        print(f"  [SKIP] No .xlsx found in {session_folder}")
        return None
    excel_file = xlsx_files[0]
    if len(xlsx_files) > 1:
        print(f"  [!] Multiple .xlsx found, using: {os.path.basename(excel_file)}")

    # 1. Videos
    print(f"\n  [1] Videos:")
    videos    = parse_video_files(session_folder)
    if not videos:
        print(f"  [SKIP] No .mp4 files in {session_folder}")
        return None
    vid_ivals = build_video_intervals(videos, ffmpeg_exe)

    # 2. EDF
    print(f"\n  [2] EDF:")
    edf_records = parse_edf_files(session_folder)

    # 3. Annotations
    print(f"\n  [3] Annotations: {os.path.basename(excel_file)}")
    all_seizures, target_seizures = parse_annotations(excel_file, FILTER_LABELS)
    print(f"      {len(all_seizures)} total seizures, "
          f"{len(target_seizures)} match filter")

    if not target_seizures:
        print("  [SKIP] No matching seizures to sample non-seizure clips for.")
        return {"ok_both": 0, "ok_vid_only": 0, "ok_eeg_only": 0,
                "skipped": 0, "failed": 0, "total": 0}

    # 4. Build blocked time map (ALL seizures + safety margin)
    blocked = build_blocked_intervals(all_seizures)
    print(f"\n  [4] Blocked intervals (seizures + {SAFETY_S}s safety margin): "
          f"{len(blocked)} merged regions")

    os.makedirs(output_folder, exist_ok=True)
    print(f"\n  [5] Output → {output_folder}")

    ok_both = ok_vid_only = ok_eeg_only = skipped = failed = 0
    log_rows = []

    for clip_n, sz in enumerate(target_seizures, start=1):
        sz_start   = sz["start"]
        sz_end     = sz["end"]
        clip_dur_s = PRE_BUFFER_S + sz["duration"] + POST_BUFFER_S

        print(f"\n  ── Clip {clip_n:02d}  matches seizure #{sz['idx']:02d} "
              f"({sz['label']}  {sz_start}  {sz['duration']:.1f}s + "
              f"{PRE_BUFFER_S}+{POST_BUFFER_S}s buffers = {clip_dur_s:.1f}s total)")

        # Find which video contains this seizure (to sample from same video only)
        source_video = None
        for v in vid_ivals:
            if v["start"] <= sz_start < v["end"]:
                source_video = v
                break

        if source_video is None:
            # Fall back: try any video that has enough free time
            candidate_videos = vid_ivals
        else:
            # Prefer the same video first, then fall back to others
            candidate_videos = [source_video] + [v for v in vid_ivals if v is not source_video]

        clip_start_dt = None
        chosen_video  = None

        for v in candidate_videos:
            free_windows = find_free_windows(v, blocked, clip_dur_s)
            if not free_windows:
                continue
            sample_start = sample_random_window(free_windows, clip_dur_s)
            if sample_start is not None:
                clip_start_dt = sample_start
                chosen_video  = v
                break

        if clip_start_dt is None:
            print(f"     [SKIP] No free window of {clip_dur_s:.1f}s found in any video.")
            log_rows.append([clip_n, sz["idx"], sz["label"], str(sz_start),
                             f"{clip_dur_s:.1f}s", "—", "—",
                             "NO_FREE_WINDOW", "—"])
            skipped += 1
            continue

        clip_end_dt = clip_start_dt + timedelta(seconds=clip_dur_s)
        seek_s      = (clip_start_dt - chosen_video["start"]).total_seconds()

        safe_label  = re.sub(r"[^\w]", "_", sz["label"])
        folder_name = (f"clip_{clip_n:02d}_vs_seizure_{sz['idx']:02d}"
                       f"_{safe_label}_{sz_start.strftime('%Y%m%d_%H%M%S')}")
        clip_folder = os.path.join(output_folder, folder_name)
        os.makedirs(clip_folder, exist_ok=True)

        print(f"     Sampled from : {chosen_video['fname']}")
        print(f"     Window       : {clip_start_dt}  →  {clip_end_dt}  "
              f"(seek={_hms(seek_s)}  dur={clip_dur_s:.1f}s)")

        vid_status = eeg_status = "—"

        # ── Cut video ──────────────────────────────────────────
        vid_out = os.path.join(clip_folder, "video.mp4")
        if os.path.exists(vid_out):
            print(f"     VIDEO : [already exists]")
            vid_status = "EXISTS"
        else:
            v_clip_dur = min(clip_dur_s, chosen_video["dur_s"] - seek_s)
            print(f"     VIDEO : seek={_hms(seek_s)}  dur={v_clip_dur:.0f}s  ", end="", flush=True)
            ok, err = cut_video_clip(ffmpeg_exe, chosen_video["fpath"],
                                     seek_s, v_clip_dur, vid_out, CUT_MODE)
            if ok and os.path.exists(vid_out):
                mb  = os.path.getsize(vid_out) / (1024 * 1024)
                fps = chosen_video["fps"]
                print(f"OK  ({mb:.1f}MB  {fps}fps)")
                vid_status = f"OK {mb:.1f}MB"
            else:
                errmsg = (err.strip().splitlines() or [""])[-1][:60]
                print(f"FAILED — {errmsg}")
                vid_status = "FAILED"

        # ── Cut EEG ────────────────────────────────────────────
        edf = find_edf(clip_start_dt, clip_end_dt, edf_records)
        if edf:
            eeg_out = os.path.join(clip_folder, "eeg.edf")
            if os.path.exists(eeg_out):
                print(f"     EEG   : [already exists]")
                eeg_status = "EXISTS"
            else:
                print(f"     EEG   : {edf['fname']}  ", end="", flush=True)
                ok, msg = cut_eeg_segment(edf, clip_start_dt, clip_end_dt, eeg_out)
                if ok:
                    print(f"OK  ({msg})")
                    eeg_status = f"OK {msg}"
                else:
                    print(f"FAILED — {msg}")
                    eeg_status = f"FAILED: {msg}"
        else:
            print(f"     EEG   : no matching EDF for sampled window")

        # ── Write info.txt ─────────────────────────────────────
        with open(os.path.join(clip_folder, "info.txt"), "w") as f:
            f.write(f"Non-seizure clip index   : {clip_n}\n")
            f.write(f"Matched seizure index    : {sz['idx']}\n")
            f.write(f"Matched seizure label    : {sz['label']}\n")
            f.write(f"Matched seizure start    : {sz_start}\n")
            f.write(f"Matched seizure end      : {sz_end}\n")
            f.write(f"Matched seizure dur (s)  : {sz['duration']:.2f}\n")
            f.write(f"Clip start               : {clip_start_dt}\n")
            f.write(f"Clip end                 : {clip_end_dt}\n")
            f.write(f"Clip duration (s)        : {clip_dur_s:.2f}\n")
            f.write(f"Pre-buffer (s)           : {PRE_BUFFER_S}\n")
            f.write(f"Post-buffer (s)          : {POST_BUFFER_S}\n")
            f.write(f"Safety margin (s)        : {SAFETY_S}\n")
            f.write(f"Video file               : {chosen_video['fname']}\n")
            f.write(f"Video seek               : {_hms(seek_s)} ({seek_s:.2f}s)\n")
            if edf:
                f.write(f"EDF file                 : {edf['fname']}\n")
                f.write(f"EEG channels             : {edf['n_ch']}\n")
                f.write(f"EEG sfreq (Hz)           : {edf['sfreq']}\n")

        v_ok = "OK" in vid_status
        e_ok = "OK" in eeg_status
        if v_ok and e_ok:   ok_both     += 1
        elif v_ok:          ok_vid_only += 1
        elif e_ok:          ok_eeg_only += 1
        else:               failed      += 1

        log_rows.append([clip_n, sz["idx"], sz["label"], str(sz_start),
                         f"{clip_dur_s:.1f}s",
                         str(clip_start_dt), str(clip_end_dt),
                         vid_status, eeg_status, folder_name])

    # Write log CSV
    log_path = os.path.join(output_folder, "_non_seizure_clip_log.csv")
    with open(log_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Clip#", "Seizure#", "Label", "Seizure Start",
                    "Clip Dur", "Clip Start", "Clip End",
                    "Video Status", "EEG Status", "Folder"])
        w.writerows(log_rows)

    print(f"\n  Log → {log_path}")
    return {"ok_both": ok_both, "ok_vid_only": ok_vid_only,
            "ok_eeg_only": ok_eeg_only, "skipped": skipped,
            "failed": failed, "total": len(target_seizures)}


# ─────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────

def main():
    print("\n" + "=" * 72)
    print("  NON-SEIZURE CLIP SAMPLER")
    print("=" * 72)
    print(f"  Pre={PRE_BUFFER_S}s | Post={POST_BUFFER_S}s | Safety={SAFETY_S}s | Mode={CUT_MODE}"
          + (f" | Filter={FILTER_LABELS}" if FILTER_LABELS else "")
          + (f" | Seed={args.seed}" if args.seed is not None else ""))

    ffmpeg_exe = check_deps()

    if args.batch:
        base = Path(args.batch)
        if not base.is_dir():
            sys.exit(f"[ERROR] --batch path does not exist: {base}")

        sessions = sorted([d for d in base.iterdir() if d.is_dir()])
        print(f"\n  Batch root : {base}")
        print(f"  Sessions   : {len(sessions)}\n")

        totals = {"ok_both": 0, "ok_vid_only": 0, "ok_eeg_only": 0,
                  "skipped": 0, "failed": 0, "total": 0}

        for i, session in enumerate(sessions, 1):
            print("\n" + "─" * 72)
            print(f"  SESSION {i}/{len(sessions)}: {session.name}")
            print("─" * 72)
            out = Path(args.output) / session.name if args.output else None
            result = process_session(str(session), ffmpeg_exe,
                                     output_folder=str(out) if out else None)
            if result:
                for k in totals:
                    totals[k] += result[k]

        print("\n" + "=" * 72)
        print("  BATCH COMPLETE — OVERALL SUMMARY")
        print("=" * 72)
        print(f"  Sessions processed : {len(sessions)}")
        print(f"  Both video + EEG   : {totals['ok_both']}")
        print(f"  Video only         : {totals['ok_vid_only']}")
        print(f"  EEG only           : {totals['ok_eeg_only']}")
        print(f"  Skipped            : {totals['skipped']}")
        print(f"  Failed             : {totals['failed']}")
        print(f"  Total clips        : {totals['total']}\n")

    else:
        session = Path(args.session)
        if not session.is_dir():
            sys.exit(f"[ERROR] --session path does not exist: {session}")
        out = args.output or str(session / "non_seizure_clips")

        result = process_session(str(session), ffmpeg_exe, output_folder=out)
        r = result or {}
        print("\n" + "=" * 72)
        print("  SUMMARY")
        print("=" * 72)
        print(f"  Both video + EEG : {r.get('ok_both', 0)}")
        print(f"  Video only       : {r.get('ok_vid_only', 0)}")
        print(f"  EEG only         : {r.get('ok_eeg_only', 0)}")
        print(f"  Skipped          : {r.get('skipped', 0)}")
        print(f"  Failed           : {r.get('failed', 0)}")
        print(f"\n  Output → {os.path.abspath(out)}\n")


if __name__ == "__main__":
    main()
